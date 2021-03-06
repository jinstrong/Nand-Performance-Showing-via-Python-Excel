    #***********************************************************************************
    #This API is for auto processing Vt loss data
    #***********************************************************************************

from Tkinter import *
import tkSimpleDialog
from tkFileDialog import *
import glob
import os
from openpyxl import Workbook
from openpyxl.drawing import Image
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt


def init():
    global row_shift,col_shift
    ws0.cell(row = row_shift, column = col_shift).value='file name'
    ws0.cell(row = row_shift, column = col_shift+1).value='page name'
    ws0.cell(row = row_shift, column = col_shift+2).value='head or tail'
    ws0.cell(row = row_shift, column = col_shift+3).value='Voltage Spec'    
    
def main(): 
    global excel_name,file,row_shift,col_shift,file_num
    VT = [[0 for x in xrange(2)] for x in xrange(2000)] 
    PG = [0 for x in xrange(20)]
    
    #***********************************************************************************
    #Searching files within specified folder
    #***********************************************************************************
    
    root = Tk()
    root.wm_title("Vt Loss Data Processing")
    w = Label(root, text="Please select the folder for the Vt Loss Data to be processed.") 
    dirname = askdirectory()
    print "Selecting folder:",dirname
    
    fold=dirname.replace('/',' ').split()
    excel_name=fold[len(fold)-1]+' Vt Loss Data.xlsx'
    print "Please close this dialog by clicking X on the right-top"
    
    w.pack()
    root.mainloop()
    os.chdir(dirname)
    
    for file in glob.glob("*.txt"):
        VT_num=-1
        file_num=file_num+1
        PG_num=-1
        line_num=0
        with open(file) as f:
            
    #***********************************************************************************
    #Copying data into VT 2-D array
    #***********************************************************************************
    
            for lines in f:
                line_num=line_num+1
                if (line_num % 122 >2):
                    VT[line_num][0]=float(lines.split()[0])
                    pop=int(lines.split()[1])
                    VT[line_num][1]=pop
        
                elif(line_num % 122 ==2):
                    PG_num=PG_num+1
                    VT[line_num][0]=lines.replace(':',' ').replace('h',' ').split()[11]
                    PG[PG_num]=VT[line_num][0]         
                        
    #***********************************************************************************
    #  1: draw diagram and save to excel file
    #  2: check head and tail, and write corresponding data into excel file
    #***********************************************************************************
   
            for i in range(0,PG_num+1):
                VT_num=VT_num+1
                HD=1
                head=0
                tail=0
                VT1=np.array(VT)[i*122+3:i*122+122]
                dia_name=file[:-4]+' page  '+PG[i]
                
                
                fig=plt.figure(figsize=(4,2), dpi=100)
                axes = fig.add_axes([0.2, 0.1, 0.8, 0.7])  # left, bottom, width, height (range 0 to 1)
                
                axes.plot(VT1[:,0],VT1[:,1],'r--')
                axes.set_xlabel('Voltage')
                axes.set_ylabel('Population')
                axes.set_title(dia_name)
                fig.savefig(dia_name+'.png',dpi=100)
                                             
                img = Image(dia_name+'.png')
                
                img.anchor(ws1.cell(row=1+10*VT_num,column=1+file_num*8))
                ws1.add_image(img)
                
                
                
                for j in range(2,120): 
                    if (HD==1):
                        if (
                            ((VT[i*122+j+1][1]<=10) and (VT[i*122+j][1]<=10) and (VT[i*122+j+2][1]>=10) and (VT[i*122+j+4][1]>10) )
                            or 
                            ((VT[i*122+j+1][1]<=10) and (VT[i*122+j+2][1]>=10)and(VT[i*122+j+3][1]>10)and (VT[i*122+j+1][0]<1.0))
                            ):
                            head=head+1
                            print 'head',i*122+j+1,VT[i*122+j+1][0],VT[i*122+j+1][1]
                            HD=0
                            ws0.cell(row = row_shift+i*6+head+tail+file_num*6*(PG_num+1), column = col_shift).value=file
                            ws0.cell(row = row_shift+i*6+head+tail+file_num*6*(PG_num+1), column = col_shift+1).value=PG[i]
                            ws0.cell(row = row_shift+i*6+head+tail+file_num*6*(PG_num+1), column = col_shift+2).value='head-'+str(head)
                            ws0.cell(row = row_shift+i*6+head+tail+file_num*6*(PG_num+1), column = col_shift+3).value=VT[i*122+j+1][0]
                    else:
                        if( (VT[i*122+j-2][1]>10)and (VT[i*122+j][1]>10) and (VT[i*122+j+1][1]<=10)and (VT[i*122+j+2][1]<=10)and (VT[i*122+j+1][0]>1.2)):
                            print 'tail',i*122+j+1,VT[i*122+j+1][0],VT[i*122+j+1][1]
                            tail=tail+1
                            HD=1
                            ws0.cell(row = row_shift+i*6+head+tail+file_num*6*(PG_num+1), column = col_shift).value=file
                            ws0.cell(row = row_shift+i*6+head+tail+file_num*6*(PG_num+1), column = col_shift+1).value=PG[i]
                            ws0.cell(row = row_shift+i*6+head+tail+file_num*6*(PG_num+1), column = col_shift+2).value='tail-'+str(tail)
                            ws0.cell(row = row_shift+i*6+head+tail+file_num*6*(PG_num+1), column = col_shift+3).value=VT[i*122+j+1][0]
                           
                        elif((VT[i*122+j][1]<50)and (VT[i*122+j][1]-VT[i*122+j+1][1]>5)and (VT[i*122+j+1][1]>10)and(VT[i*122+j+1][1]-VT[i*122+j+2][1]<-5)and (VT[i*122+j+2][1]<50)):
                            if(VT[i*122+j][0]>1.5):
                                print 'tail and head',VT[i*122+j+1][0],VT[i*122+j+1][1]
                            
                                tail=tail+1
                                ws0.cell(row = row_shift+i*6+head+tail+file_num*6*(PG_num+1), column = col_shift).value=file
                                ws0.cell(row = row_shift+i*6+head+tail+file_num*6*(PG_num+1), column = col_shift+1).value=PG[i]
                                ws0.cell(row = row_shift+i*6+head+tail+file_num*6*(PG_num+1), column = col_shift+2).value='tail-'+str(tail)
                                ws0.cell(row = row_shift+i*6+head+tail+file_num*6*(PG_num+1), column = col_shift+3).value=VT[i*122+j+1][0]  
                                head=head+1
                                ws0.cell(row = row_shift+i*6+head+tail+file_num*6*(PG_num+1), column = col_shift).value=file
                                ws0.cell(row = row_shift+i*6+head+tail+file_num*6*(PG_num+1), column = col_shift+1).value=PG[i]
                                ws0.cell(row = row_shift+i*6+head+tail+file_num*6*(PG_num+1), column = col_shift+2).value='head-'+str(head)
                                ws0.cell(row = row_shift+i*6+head+tail+file_num*6*(PG_num+1), column = col_shift+3).value=VT[i*122+j+1][0]                      
                        
                print 'head',head,'tail',tail            

if __name__ == '__main__':
    
    #***********************************************************************************
    #Excel format specifying
    #***********************************************************************************
    row_shift=9
    col_shift=3
    file_num=-1
    excel_name=' '
    wb=Workbook()
    ws = wb.active
    ws.title="Summary"
    ws0 = wb.create_sheet()
    ws0.title="Vt Loss"
    ws1 = wb.create_sheet()
    ws1.title="Diagram"
    ws2 = wb.create_sheet()
    ws2.title="read"
    ws3 = wb.create_sheet()
    ws3.title="FBC"
    init()
    main()
    wb.save(excel_name)
    
